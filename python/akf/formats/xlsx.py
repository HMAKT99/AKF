"""AKF v1.0 — XLSX format handler.

Embeds AKF metadata into Excel XLSX files using the OOXML Custom XML Part
mechanism. Basic embed/extract operations require NO external dependencies
(uses stdlib zipfile only). Advanced operations like auto_enrich and create
optionally use openpyxl for cell-level introspection.

Usage:
    from akf.formats.xlsx import embed, extract, is_enriched, scan

    embed("data.xlsx", {"akf": "1.0", "claims": [...]})
    meta = extract("data.xlsx")
"""

from typing import List, Optional

from .base import AKFFormatHandler, ScanReport


class XLSXHandler(AKFFormatHandler):
    """Handler for Microsoft Excel XLSX files."""

    FORMAT_NAME: str = "XLSX"
    EXTENSIONS: List[str] = [".xlsx"]
    MODE: str = "embedded"
    MECHANISM: str = "OOXML Custom XML Part"
    DEPENDENCIES: List[str] = []  # Basic ops use zipfile only

    def embed(self, filepath: str, metadata: dict) -> None:
        """Embed AKF metadata into an XLSX file.

        Uses stdlib zipfile to add a custom XML part to the OOXML archive.
        No openpyxl dependency required.
        """
        from ._ooxml import embed_in_ooxml

        embed_in_ooxml(filepath, metadata)

    def extract(self, filepath: str) -> Optional[dict]:
        """Extract AKF metadata from an XLSX file.

        Returns None if no AKF metadata is embedded.
        """
        from ._ooxml import extract_from_ooxml

        return extract_from_ooxml(filepath)

    def is_enriched(self, filepath: str) -> bool:
        """Check if an XLSX file has AKF metadata embedded."""
        from ._ooxml import is_ooxml_enriched

        return is_ooxml_enriched(filepath)

    def auto_enrich(
        self,
        filepath: str,
        agent_id: str,
        default_tier: int = 3,
        classification: Optional[str] = None,
    ) -> None:
        """Auto-enrich an XLSX file with AKF metadata.

        If openpyxl is available, reads cell values from the active sheet
        and creates per-row claims. Otherwise, embeds basic metadata only.
        """
        meta = self._build_auto_metadata(filepath, agent_id, default_tier, classification)

        # Try to extract cell data for claims if openpyxl is available
        try:
            from openpyxl import load_workbook

            wb = load_workbook(filepath, read_only=True, data_only=True)
            ws = wb.active
            claims = []
            if ws is not None:
                for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
                    # Build a text representation of the row
                    values = [str(v) for v in row if v is not None]
                    if values:
                        text = ", ".join(values)
                        if len(text) > 10:
                            claims.append(
                                {
                                    "location": "sheet:{},row:{}".format(
                                        ws.title, row_idx + 1
                                    ),
                                    "c": text[:200],
                                    "t": 0.7,
                                    "src": agent_id,
                                    "ai": True,
                                    "tier": default_tier,
                                }
                            )
            wb.close()
            if claims:
                meta["claims"] = claims
                meta["ai_contribution"] = 1.0
        except ImportError:
            pass  # No openpyxl — embed basic metadata only

        self.embed(filepath, meta)


# ---------------------------------------------------------------------------
# Module-level convenience API
# ---------------------------------------------------------------------------

_handler = XLSXHandler()
embed = _handler.embed
extract = _handler.extract
is_enriched = _handler.is_enriched
scan = _handler.scan
auto_enrich = _handler.auto_enrich


def create(
    output: str,
    data: list,
    sheet_name: str = "Sheet1",
    classification: Optional[str] = None,
    author: Optional[str] = None,
) -> None:
    """Create a new XLSX with AKF metadata from structured data.

    Requires openpyxl. Install with: pip install akf[xlsx]

    Args:
        output: Output file path.
        data: List of rows. Each row is a list of cell values, or a dict with:
            - row: List of cell values
            - akf: (optional) Dict with claim metadata (t, src, ai, etc.)
        sheet_name: Name of the worksheet.
        classification: Optional security classification label.
        author: Optional author/agent identifier.

    Example::

        create("data.xlsx", [
            {"row": ["Company", "Revenue", "Trust"], "akf": {"t": 0.95}},
            {"row": ["Acme Corp", 4200000, 0.98]},
        ])
    """
    try:
        from openpyxl import Workbook
    except ImportError:
        raise ImportError(
            "Creating XLSX requires openpyxl. Install with: pip install akf[xlsx]"
        )

    import hashlib
    from datetime import datetime, timezone

    wb = Workbook()
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet(sheet_name)
    else:
        ws.title = sheet_name

    claims = []

    for i, item in enumerate(data):
        if isinstance(item, dict):
            row_data = item.get("row", [])
            akf_meta = item.get("akf")
        else:
            row_data = item
            akf_meta = None

        ws.append(row_data)

        if akf_meta is not None:
            text = ", ".join(str(v) for v in row_data if v is not None)
            claim = {
                "location": "sheet:{},row:{}".format(sheet_name, i + 1),
                "c": text[:200],
            }
            claim.update(akf_meta)
            claims.append(claim)

    wb.save(output)

    # Compute hash of the saved file
    now = datetime.now(timezone.utc).isoformat()
    with open(output, "rb") as f:
        file_hash = "sha256:" + hashlib.sha256(f.read()).hexdigest()

    # Build metadata
    metadata = {
        "akf": "1.0",
        "generated_at": now,
        "overall_trust": (
            sum(c.get("t", 0.5) for c in claims) / len(claims) if claims else 0.5
        ),
        "ai_contribution": (
            sum(1 for c in claims if c.get("ai")) / len(claims) if claims else 0.0
        ),
        "claims": claims,
        "provenance": [
            {
                "actor": author or "unknown",
                "action": "created",
                "at": now,
                "hash": file_hash,
            }
        ],
        "integrity_hash": file_hash,
    }

    if classification is not None:
        metadata["classification"] = classification

    embed(output, metadata)
